# 新建xls文档 A，包含sheet:"result"
# 准备写入数据

# 读取文件夹B 中的所有xls文档

# for 循环B中的每个文档
## 提取文档名称 name
## 读取文档第一行的所有单元格的值和批注
## 将名称和第一行的值和批注写入A的sheet:"result"中，每个文档一行

# 保存A
import os
import openpyxl
import xlrd

# 新建xls文档 A，包含sheet:"result"
wb_A = openpyxl.Workbook()
sheet_A = wb_A.active
sheet_A.title = "result"

# 准备写入数据
sheet_A.append(["Document Name", "Cell Values", "Comments"])

# 读取文件夹B 中的所有xls文档
folder_B = "d:/ABecky/code/AA"
for filename in os.listdir(folder_B):
    if filename.endswith(".xls"):
        filepath = os.path.join(folder_B, filename)
        
        # 提取文档名称 name
        name = os.path.splitext(filename)[0]
        
        # 读取文档第一行的所有单元格的值和批注
        wb_B = xlrd.open_workbook(filepath)
        sheet_B = wb_B.sheet_by_index(0)
        first_row_values = [sheet_B.cell_value(0, col) for col in range(sheet_B.ncols)]
        # first_row_comments = [sheet_B.cell(0, col).comment.text if sheet_B.cell(0, col).comment else "" for col in range(sheet_B.ncols)]
        
        # 将名称和第一行的值和批注写入A的sheet:"result"中，每个文档一行
        sheet_A.append([name, str(first_row_values), "0"])

# 保存A
wb_A.save("result.xlsx")